"""
Excel文件处理器
支持多种Excel格式和多工作表处理
"""
import pandas as pd
import openpyxl
import xlrd
import zipfile
import re
from io import BytesIO
from typing import Dict, List, Optional, Tuple, Any
import logging

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Excel文件处理器"""

    # 支持的文件扩展名
    SUPPORTED_EXTENSIONS = {
        # Excel格式
        '.xlsx': 'Excel 2007+ (OpenXML)',
        '.xls': 'Excel 97-2003',
        '.xlsm': 'Excel带宏工作簿',
        '.xlsb': 'Excel二进制工作簿',
        '.xltx': 'Excel模板',
        '.xltm': 'Excel宏模板',
        # CSV格式
        '.csv': '逗号分隔值',
        '.tsv': '制表符分隔值',
        '.txt': '文本文件',
    }

    @staticmethod
    def get_worksheets(file_bytes: bytes, file_ext: str) -> List[Dict[str, Any]]:
        """
        获取Excel文件中的所有工作表信息

        Args:
            file_bytes: 文件字节数据
            file_ext: 文件扩展名

        Returns:
            工作表信息列表
            [
                {
                    'name': 'Sheet1',
                    'index': 0,
                    'rows': 100,
                    'has_content': True  # 新增字段，表示是否包含有效内容
                },
                ...
            ]
        """
        try:
            if file_ext in ['.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']:
                # 使用openpyxl处理新格式
                try:
                    workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                except ValueError as e:
                    # 检查是否是样式表错误
                    if 'stylesheet' in str(e) or 'vertAlign' in str(e):
                        logger.warning(f"get_worksheets: Excel文件样式表错误，尝试自动修复: {e}")
                        # 修复文件
                        repaired_bytes = ExcelHandler._repair_xlsx_file(file_bytes)
                        workbook = openpyxl.load_workbook(BytesIO(repaired_bytes), read_only=True, data_only=True)
                    else:
                        raise
                        
                worksheets = []
                for idx, sheet_name in enumerate(workbook.sheetnames):
                    sheet = workbook[sheet_name]
                    rows = sheet.max_row if sheet.max_row else 0
                    # 检查工作表是否包含有效内容
                    has_content = ExcelHandler._has_meaningful_content(sheet, file_ext)
                    worksheets.append({
                        'name': sheet_name,
                        'index': idx,
                        'rows': rows,
                        'has_content': has_content
                    })
                workbook.close()
                return worksheets

            elif file_ext == '.xls':
                # 使用xlrd处理旧格式
                workbook = xlrd.open_workbook(file_contents=file_bytes)
                worksheets = []
                for idx, sheet_name in enumerate(workbook.sheet_names()):
                    sheet = workbook.sheet_by_index(idx)
                    # 检查工作表是否包含有效内容
                    has_content = ExcelHandler._has_meaningful_content_xls(sheet)
                    worksheets.append({
                        'name': sheet_name,
                        'index': idx,
                        'rows': sheet.nrows,
                        'has_content': has_content
                    })
                return worksheets

            elif file_ext in ['.csv', '.tsv', '.txt']:
                # CSV文件只有一个"工作表"
                # 先读取确定行数
                separator = '\t' if file_ext == '.tsv' else ','
                df = pd.read_csv(BytesIO(file_bytes), sep=separator, nrows=1000)
                return [{
                    'name': 'Data',
                    'index': 0,
                    'rows': len(df),
                    'has_content': len(df) > 0
                }]

            else:
                raise ValueError(f"不支持的文件格式: {file_ext}")

        except Exception as e:
            logger.error(f"获取工作表信息失败: {e}", exc_info=True)
            raise

    @staticmethod
    def _has_meaningful_content(sheet, file_ext: str) -> bool:
        """
        检查工作表是否包含有意义的内容
        """
        if sheet.max_row == 0:
            return False
            
        # 读取前几行数据检查是否包含有意义的内容
        content_cells = 0
        max_check_rows = min(10, sheet.max_row)
        
        for row_idx in range(1, max_check_rows + 1):
            for col_idx in range(1, min(10, sheet.max_column + 1)):  # 检查前10列
                try:
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip() != '':
                        content_cells += 1
                        if content_cells >= 3:  # 如果找到至少3个非空单元格，认为有内容
                            return True
                except:
                    continue
        
        return content_cells > 0

    @staticmethod
    def _has_meaningful_content_xls(sheet) -> bool:
        """
        检查xls格式工作表是否包含有意义的内容
        """
        if sheet.nrows == 0:
            return False
            
        # 读取前几行数据检查是否包含有意义的内容
        content_cells = 0
        max_check_rows = min(10, sheet.nrows)
        
        for row_idx in range(max_check_rows):
            for col_idx in range(min(10, sheet.ncols)):  # 检查前10列
                try:
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None and str(cell_value).strip() != '':
                        content_cells += 1
                        if content_cells >= 3:  # 如果找到至少3个非空单元格，认为有内容
                            return True
                except:
                    continue
        
        return content_cells > 0

    @staticmethod
    def find_first_content_sheet(file_bytes: bytes, file_ext: str) -> Optional[Dict[str, Any]]:
        """
        查找第一个包含内容的工作表
        """
        worksheets = ExcelHandler.get_worksheets(file_bytes, file_ext)
        for sheet in worksheets:
            if sheet.get('has_content', True):  # 如果没有has_content字段，默认认为有内容
                return sheet
        return worksheets[0] if worksheets else None  # 如果都没有内容，返回第一个

    @staticmethod
    def read_sheet(file_bytes: bytes, file_ext: str, sheet_index: int = 0,
                  sheet_name: Optional[str] = None, **kwargs) -> pd.DataFrame:
        """
        读取指定工作表

        Args:
            file_bytes: 文件字节数据
            file_ext: 文件扩展名
            sheet_index: 工作表索引（默认0）
            sheet_name: 工作表名称（优先使用）
            **kwargs: 传递给pandas.read_excel/read_csv的其他参数

        Returns:
            DataFrame对象
        """
        try:
            if file_ext in ['.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']:
                # 新格式Excel
                engine = 'openpyxl' if file_ext in ['.xlsx', '.xlsm', '.xltx', '.xltm'] else 'pyxlsb'
                read_kwargs = {
                    'dtype': str,
                    'header': None,
                }
                read_kwargs.update(kwargs)

                if sheet_name:
                    read_kwargs['sheet_name'] = sheet_name
                else:
                    read_kwargs['sheet_name'] = sheet_index

                try:
                    df = pd.read_excel(BytesIO(file_bytes), engine=engine, **read_kwargs)
                    return df
                except (ValueError, Exception) as e:
                    # 检查是否是样式表错误
                    error_str = str(e)
                    if 'stylesheet' in error_str or 'vertAlign' in error_str or 'XML' in error_str or 'could not read' in error_str:
                        logger.warning(f"Excel文件样式表错误，尝试自动修复: {e}")
                        df = ExcelHandler._repair_and_read_xlsx(file_bytes, read_kwargs)
                        return df
                    else:
                        raise

            elif file_ext == '.xls':
                # 旧格式Excel
                read_kwargs = {
                    'dtype': str,
                    'header': None,
                    'engine': 'xlrd'
                }
                read_kwargs.update(kwargs)

                if sheet_name:
                    read_kwargs['sheet_name'] = sheet_name
                else:
                    read_kwargs['sheet_name'] = sheet_index

                df = pd.read_excel(BytesIO(file_bytes), **read_kwargs)
                return df

            elif file_ext in ['.csv', '.tsv', '.txt']:
                # CSV/TSV文件
                separator = '\t' if file_ext == '.tsv' else ','
                read_kwargs = {
                    'sep': separator,
                    'dtype': str,
                    'header': None,
                }
                read_kwargs.update(kwargs)

                df = pd.read_csv(BytesIO(file_bytes), **read_kwargs)
                return df

            else:
                raise ValueError(f"不支持的文件格式: {file_ext}")

        except Exception as e:
            logger.error(f"读取工作表失败: {e}", exc_info=True)
            raise

    @staticmethod
    def detect_csv_separator(file_bytes: bytes, sample_size: int = 1024) -> str:
        """
        自动检测CSV分隔符

        Args:
            file_bytes: 文件字节数据
            sample_size: 采样大小

        Returns:
            检测到的分隔符
        """
        import re

        # 读取样本
        sample = file_bytes[:sample_size].decode('utf-8', errors='ignore')

        # 统计常见分隔符出现的次数
        separators = {
            ',': sample.count(','),
            ';': sample.count(';'),
            '\t': sample.count('\t'),
            '|': sample.count('|'),
        }

        # 返回出现次数最多的分隔符
        return max(separators.items(), key=lambda x: x[1])[0]

    @staticmethod
    def _repair_xlsx_file(file_bytes: bytes) -> bytes:
        """
        修复损坏的Excel文件并返回修复后的字节数据

        修复常见的样式表错误，如无效的vertAlign值、无效的XML属性等

        Args:
            file_bytes: 文件字节数据

        Returns:
            修复后的文件字节数据
        """
        import re
        from io import BytesIO

        # Excel文件实际上是一个ZIP文件
        zip_buffer = BytesIO(file_bytes)

        # 读取ZIP文件
        with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
            # 尝试读取并修复样式文件
            try:
                styles_content = zip_ref.read('xl/styles.xml')
                
                # 修复1: 查找并替换无效的vertAlign值
                valid_values = {'baseline', 'superscript', 'subscript'}
                fixed_content = styles_content.decode('utf-8', errors='ignore')
                
                pattern = r'<vertAlign val="([^"]+)"/>'
                matches = re.findall(pattern, fixed_content)
                
                for match in matches:
                    if match not in valid_values:
                        fixed_content = fixed_content.replace(f'val="{match}"', 'val="baseline"')
                        logger.info(f"修复无效的vertAlign值: {match} -> baseline")
                
                # 修复2: 处理其他可能的无效属性值
                # 修复无效的u属性（下划线样式）
                u_pattern = r'<u val="([^"]+)"/>'
                u_valid_values = {'none', 'single', 'double', 'singleAccounting', 'doubleAccounting'}
                u_matches = re.findall(u_pattern, fixed_content)
                for match in u_matches:
                    if match not in u_valid_values:
                        fixed_content = fixed_content.replace(f'<u val="{match}"/>', '<u val="none"/>')
                        logger.info(f"修复无效的u值: {match} -> none")
                
                # 修复3: 移除可能引起问题的空属性
                fixed_content = re.sub(r'<(\w+)\s+/>', r'<\1/>', fixed_content)
                
                # 修复4: 确保XML声明存在
                if not fixed_content.startswith('<?xml'):
                    fixed_content = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + fixed_content
                    logger.info("添加XML声明")
                
                # 修复5: 处理可能损坏的font标签
                fixed_content = re.sub(r'<font[^>]*\s+\s*>', '<font>', fixed_content)
                
                # 修复6: 移除可能存在的BOM标记
                if fixed_content.startswith('\ufeff'):
                    fixed_content = fixed_content[1:]
                    logger.info("移除BOM标记")
                
                # 创建新的ZIP文件
                new_zip_buffer = BytesIO()
                with zipfile.ZipFile(new_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                    for name in zip_ref.namelist():
                        if name == 'xl/styles.xml':
                            new_zip.writestr(name, fixed_content.encode('utf-8'))
                        else:
                            new_zip.writestr(name, zip_ref.read(name))
                
                # 返回修复后的字节数据
                repaired_bytes = new_zip_buffer.getvalue()
                logger.info("成功修复Excel文件")
                return repaired_bytes
                
            except Exception as e:
                logger.error(f"修复Excel文件失败: {e}")
                raise

    @staticmethod
    def _repair_and_read_xlsx(file_bytes: bytes, read_kwargs: dict) -> pd.DataFrame:
        """
        修复损坏的Excel文件并读取

        修复常见的样式表错误，如无效的vertAlign值、无效的XML属性等

        Args:
            file_bytes: 文件字节数据
            read_kwargs: pandas.read_excel的参数

        Returns:
            DataFrame对象
        """
        # 使用 _repair_xlsx_file 方法修复文件
        repaired_bytes = ExcelHandler._repair_xlsx_file(file_bytes)
        
        # 读取修复后的文件
        df = pd.read_excel(BytesIO(repaired_bytes), engine='openpyxl', **read_kwargs)
        logger.info("成功读取修复后的Excel文件")
        return df

    @staticmethod
    def get_supported_formats() -> Dict[str, str]:
        """
        获取支持的文件格式列表

        Returns:
            格式字典 {扩展名: 描述}
        """
        return ExcelHandler.SUPPORTED_EXTENSIONS

    @staticmethod
    def is_supported(file_ext: str) -> bool:
        """
        检查文件格式是否支持

        Args:
            file_ext: 文件扩展名

        Returns:
            是否支持
        """
        return file_ext.lower() in ExcelHandler.SUPPORTED_EXTENSIONS